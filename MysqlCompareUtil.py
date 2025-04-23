import pymysql
import openpyxl
from typing import Dict, List, Tuple

def get_table_structure(conn_config: Dict) -> Dict[str, List[Tuple]]:
    """
    获取数据库表结构信息
    :param conn_config: 数据库连接配置
    :return: 字典{表名: [字段信息]}
    """
    connection = pymysql.connect(**conn_config)
    cursor = connection.cursor()
    
    # 获取所有表名
    cursor.execute("SHOW TABLES")
    tables = [table[0] for table in cursor.fetchall()]
    
    table_info = {}
    for table in tables:
        # 获取表结构
        cursor.execute(f"SHOW CREATE TABLE {table}")
        create_table_sql = cursor.fetchone()[1]
        
        # 获取字段信息
        cursor.execute(f"DESCRIBE {table}")
        fields = cursor.fetchall()
        
        table_info[table] = fields
    
    cursor.close()
    connection.close()
    return table_info

def compare_table_structures(db1: Dict, db2: Dict) -> Dict:
    """
    比较两个数据库的表结构差异
    :param db1: 第一个数据库的表结构
    :param db2: 第二个数据库的表结构
    :return: 差异结果
    """
    result = {
        'only_in_sit': [],    # 仅存在于SIT数据库中的表名列表
        'only_in_uat': [],    # 仅存在于UAT数据库中的表名列表
        'different_tables': {} # 两个数据库共有的表但结构有差异的表
    }
    
    # 找出只在其中一个数据库中存在的表
    tables_db1 = set(db1.keys())
    tables_db2 = set(db2.keys())
    
    result['only_in_sit'] = list(tables_db1 - tables_db2)
    result['only_in_uat'] = list(tables_db2 - tables_db1)
    
    # 比较共有表的结构差异
    common_tables = tables_db1 & tables_db2
    for table in common_tables:
        fields1 = {field[0]: field for field in db1[table]}
        fields2 = {field[0]: field for field in db2[table]}
        
        diff_fields = {
            'only_in_sit': [],    # 仅存在于SIT数据库中的字段名列表
            'only_in_uat': [],    # 仅存在于UAT数据库中的字段名列表
            'different_fields': {} # 两个数据库共有的字段但属性有差异的字段
        }
        
        # 比较字段差异
        field_names1 = set(fields1.keys())
        field_names2 = set(fields2.keys())
        
        diff_fields['only_in_sit'] = list(field_names1 - field_names2)
        diff_fields['only_in_uat'] = list(field_names2 - field_names1)
        
        # 比较共有字段的属性差异
        common_fields = field_names1 & field_names2
        for field in common_fields:
            if fields1[field] != fields2[field]:
                diff_fields['different_fields'][field] = {
                    'db1': fields1[field],  # SIT数据库中该字段的完整属性
                    'db2': fields2[field],  # UAT数据库中该字段的完整属性
                    'default_value_diff': fields1[field][4] != fields2[field][4]  # 比较默认值差异
                }
        
        if any(diff_fields.values()):
            result['different_tables'][table] = diff_fields
    
    return result

def print_comparison_result(result: Dict):
    """
    打印比较结果
    :param result: 比较结果
    """
    print("\n===== 表结构差异比较结果 =====")
    
    if result['only_in_sit']:
        print("\n仅存在于SIT数据库的表:")
        print("# 这些表在SIT数据库中存在但在UAT数据库中不存在")
        for table in result['only_in_sit']:
            print(f"  - {table}")
    
    if result['only_in_uat']:
        print("\n仅存在于UAT数据库的表:")
        print("# 这些表在UAT数据库中存在但在SIT数据库中不存在")
        for table in result['only_in_uat']:
            print(f"  - {table}")
    
    if result['different_tables']:
        print("\n表结构差异:")
        print("# 这些表在两个数据库中都存在，但表结构有差异")
        for table, diff in result['different_tables'].items():
            print(f"\n表: {table}")
            
            if diff['only_in_sit']:
                print("  仅存在于SIT数据库的字段:")
                print("  # 这些字段只在SIT数据库的该表中存在")
                for field in diff['only_in_sit']:
                    print(f"    - {field}")
            
            if diff['only_in_uat']:
                print("  仅存在于UAT数据库的字段:")
                print("  # 这些字段只在UAT数据库的该表中存在")
                for field in diff['only_in_uat']:
                    print(f"    - {field}")
            
            if diff['different_fields']:
                print("  字段属性差异:")
                print("  # 这些字段在两个数据库中都存在，但属性有差异")
                for field, details in diff['different_fields'].items():
                    print(f"    字段名称: {field}")
                    print(f"      SIT数据库属性: 类型={details['db1'][1]}, 是否为空={details['db1'][2]}, 主键类型={details['db1'][3]}, 默认值='{details['db1'][4]}', 额外属性={details['db1'][5]}")
                    print(f"      UAT数据库属性: 类型={details['db2'][1]}, 是否为空={details['db2'][2]}, 主键类型={details['db2'][3]}, 默认值='{details['db2'][4]}', 额外属性={details['db2'][5]}")
                    if 'default_value_diff' in details and details['default_value_diff']:
                        print(f"      默认值差异: SIT数据库为'{details['db1'][4]}', UAT数据库为'{details['db2'][4]}'")
    
    print("\n===== 比较完成 =====")
    
    # 导出结果到Excel
    export_to_excel(result, "mysql_comparison_result.xlsx")

def export_to_excel(result: Dict, filename: str):
    """
    将比较结果导出到Excel文件
    :param result: 比较结果
    :param filename: 输出文件名
    """
    from datetime import datetime
    # 获取当前时间并格式化为年月日时分秒
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # 在原文件名中添加时间戳
    filename = f"{filename.split('.')[0]}_{timestamp}.xlsx"
    
    wb = openpyxl.Workbook()
    
    # 创建仅存在于SIT数据库的表工作表
    if result['only_in_sit']:
        ws = wb.create_sheet("仅存在于SIT的表")
        ws.append(["表名"])
        for table in result['only_in_sit']:
            ws.append([table])
    
    # 创建仅存在于UAT数据库的表工作表
    if result['only_in_uat']:
        ws = wb.create_sheet("仅存在于UAT的表")
        ws.append(["表名"])
        for table in result['only_in_uat']:
            ws.append([table])
    
    # 创建表结构差异工作表
    if result['different_tables']:
        ws = wb.create_sheet("表结构差异")
        ws.append(["表名", "差异类型", "字段名", "SIT数据库属性", "UAT数据库属性", "类型差异", "是否为空差异", "主键类型差异", "默认值差异", "额外属性差异"])
        
        for table, diff in result['different_tables'].items():
            # 仅存在于SIT的字段
            for field in diff['only_in_sit']:
                ws.append([table, "仅存在于SIT的字段", field, "-", "-"])
            
            # 仅存在于UAT的字段
            for field in diff['only_in_uat']:
                ws.append([table, "仅存在于UAT的字段", field, "-", "-"])
            
            # 字段属性差异
            for field, details in diff['different_fields'].items():
                db1_attrs = f"类型={details['db1'][1]}, 是否为空={details['db1'][2]}, 主键类型={details['db1'][3]}, 默认值='{details['db1'][4]}', 额外属性={details['db1'][5]}"
                db2_attrs = f"类型={details['db2'][1]}, 是否为空={details['db2'][2]}, 主键类型={details['db2'][3]}, 默认值='{details['db2'][4]}', 额外属性={details['db2'][5]}"
                type_diff = "是" if details['db1'][1] != details['db2'][1] else "否"
                null_diff = "是" if details['db1'][2] != details['db2'][2] else "否"
                key_diff = "是" if details['db1'][3] != details['db2'][3] else "否"
                default_diff = "是" if details['db1'][4] != details['db2'][4] else "否"
                extra_diff = "是" if details['db1'][5] != details['db2'][5] else "否"
                ws.append([table, "字段属性差异", field, db1_attrs, db2_attrs, type_diff, null_diff, key_diff, default_diff, extra_diff])
    
    # 删除默认创建的空工作表
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    wb.save(filename)
    print(f"\n比较结果已导出到文件: {filename}")

if __name__ == "__main__":
    # 示例用法
    # db1_config = {
    #     'host': '127.0.0.1',
    #     'port': 3306,
    #     'user': 'root',
    #     'password': '123456',
    #     'db': 'vkey1',
    #     'charset': 'utf8mb4'
    # }
    
    # db2_config = {
    #     'host': '127.0.0.1',
    #     'port': 3306,
    #     'user': 'root',
    #     'password': '123456',
    #     'db': 'vkey',
    #     'charset': 'utf8mb4'
    # }
    # db1_config = {
    #     'host': '192.168.83.117',
    #     'port': 8881,
    #     'user': 'd1cdns',
    #     'password': 'd1cdns',
    #     'db': 'd1cdns',
    #     'charset': 'utf8mb4'
    # }
    
    # db2_config = {
    #     'host': '192.168.83.117',
    #     'port': 8881,
    #     'user': 'd2cdns',
    #     'password': 'd2cdns',
    #     'db': 'd2cdns',
    #     'charset': 'utf8mb4'
    # }
    db1_config = {
        'host': '192.168.83.117',
        'port': 8881,
        'user': 'd2cdns',
        'password': 'd2cdns',
        'db': 'd2cdns',
        'charset': 'utf8mb4'
    }
    
    db2_config = {
        'host': '192.168.83.117',
        'port': 8881,
        'user': 'd3cdns',
        'password': 'd3cdns',
        'db': 'd3cdns',
        'charset': 'utf8mb4'
    }
    # 获取表结构
    db1_tables = get_table_structure(db1_config)
    db2_tables = get_table_structure(db2_config)
    
    # 比较表结构
    comparison_result = compare_table_structures(db1_tables, db2_tables)
    
    # 输出结果
    print_comparison_result(comparison_result)